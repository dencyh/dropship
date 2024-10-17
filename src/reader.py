from typing import Type, List
from openpyxl import load_workbook
import os
from models import Schedule, LogisticRequest
from config import Config

cfg = Config()
input_folder = cfg.get('input_folder')
order_file_prefix = cfg.get('order_file_prefix')
schedule_file_prefix = cfg.get('schedule_file_prefix')
base_path = cfg.get('base_path')


def get_schedules() -> List[Schedule]:
    return  extract_data_from_xlsx(find_file_by_prefix(input_folder, schedule_file_prefix), Schedule)

def get_logistic_requests() -> List[LogisticRequest]:
    return extract_data_from_xlsx(find_file_by_prefix(input_folder, order_file_prefix), LogisticRequest)

def extract_data_from_xlsx(file_path: str, data_class: Type) -> List:
    """
    Extracts data from an Excel file and returns a list of objects based on the provided data class.
    PS: Спасибо GPT за эту прекрасную функцию и чудесный комментарий

    Args:
        file_path (str): Path to the Excel file.
        data_class (Type): The class type that the data should be converted into.

    Returns:
        List: A list of instances of the specified class, filled with data from the Excel rows.
    """
    # Open the workbook and select the active worksheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    data_objects = []

    # Assuming the first row contains headers and data starts from the second row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_object = data_class(*row)  # Dynamically create an instance of the class
        data_objects.append(data_object)

    return data_objects


def find_file_by_prefix(directory, prefix):
    # List all files in the directory
    directory = os.path.join(base_path, directory)
    for filename in os.listdir(directory):
        if filename.startswith(prefix):
            # returns full path to the file
            return os.path.join(directory, filename)
    # if file not found
    raise FileNotFoundError(f'Не найден файл, который начинается на {prefix} в папке ${directory}')
