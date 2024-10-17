from typing import List

from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from models import CourierRoute, Direction


def generate_excel(courier_routes: List[CourierRoute], file_name: str):
    # Create a new workbook and sheets
    workbook = Workbook()
    sheets = {
        "Прямой поток": workbook.active,
        "Возвратный поток": workbook.create_sheet(title="Возвратный поток"),
        "Дропнуло": workbook.create_sheet(title="Дропнуло"),
        "Разобрать проблемы": workbook.create_sheet(title="Разобрать проблемы"),
    }
    workbook.active.title = "Прямой поток"

    # Set headers for the first two sheets
    headers = ["Парнер", "ФИО", "ID курьера", "Кол-во адресов"]
    for i in range(1, 11):
        headers.append(f'Адрес {i}')
        headers.append('#')

    for sheet in sheets.values():
        sheet.append(headers)

    # Lists for collecting information for the "Дропнуло" and "Разобрать проблемы" sheets
    dropped_routes = []
    problem_routes = []

    # Сортируем
    courier_routes = sorted(
        courier_routes,
        key=lambda r: (r.company, r.courier_name)
    )
    for route in courier_routes:
        drop_ship_count = sum(1 for addr in route.addresses if addr.direction == Direction.DROP_SHIP)
        return_count = sum(1 for addr in route.addresses if addr.direction == Direction.RETURN)

        if drop_ship_count > 0:
            # Add data to "Прямой поток"
            address_list = []
            for addr in route.addresses:
                if addr.direction == Direction.DROP_SHIP:
                    address_list.extend([addr.address_string, ""])
            row = [route.company, route.courier_name, route.courier_id, drop_ship_count]
            row.extend(address_list)
            sheets["Прямой поток"].append(row)

        if return_count > 0:
            # Add data to "Возвратный поток"
            address_list = []
            for addr in route.addresses:
                if addr.direction == Direction.RETURN:
                    address_list.extend([addr.address_string, ""])
            row = [route.company, route.courier_name, route.courier_id, return_count]
            row.extend(address_list)
            sheets["Возвратный поток"].append(row)

        if len(route.addresses) == 0:
            # Add data to "Дропнуло"
            dropped_routes.append([route.company, route.courier_name, route.courier_id])

        # Check for addresses with direction None
        if any(addr.direction is None for addr in route.addresses):
            # для плохих адресов собираем парами "адрес" и "статус заказа"
            bad_addresses = filter(lambda r: r.direction is None, route.addresses)
            # создаем развернутый список, чтобы запушить в строку
            flattened_list = [item for addr in bad_addresses for item in [addr.address_string, addr.status]]
            row = [route.company, route.courier_name, route.courier_id, len(flattened_list)/2]
            row.extend(flattened_list)
            problem_routes.append(row)

    # Populate the "Дропнуло" sheet
    for route in dropped_routes:
        sheets["Дропнуло"].append(route)

    # Populate the "Разобрать проблемы" sheet
    for route in problem_routes:
        sheets["Разобрать проблемы"].append(route)


    # Стили
    for sheet in sheets.values():
        set_column_widths(sheet)
    # Save the workbook
    workbook.save(file_name)

def set_column_widths(sheet):
    # Define the widths for the first four columns
    column_widths = {
        'A': 20,  # First column
        'B': 30,  # Second column
        'C': 8,   # Third column
        'D': 8,    # Fourth column
    }

    # Apply the widths for the first four columns
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    # Apply the pattern for the rest of the columns
    for i in range(5, 27):  # Columns E (5) to Z (26)
        column_letter = chr(64 + i)  # Convert number to letter (E=5, F=6, ..., Z=26)
        if (i - 1) % 2 == 0:  # 150px for columns E, H, K, ...
            sheet.column_dimensions[column_letter].width = 25
        elif (i - 1) % 2 == 1:  # 25px for columns F, I, L, ...
            sheet.column_dimensions[column_letter].width = 8

    for column_cells in sheet.columns:
        for cell in column_cells:
            cell.alignment = Alignment(wrap_text=True)